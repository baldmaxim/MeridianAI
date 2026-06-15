"""Обработка документов (Этап 4): скачать из S3 → извлечь текст по сегментам
(страницы PDF / листы XLSX) → разбить на чанки → сохранить DocumentChunk.

Логируем только метаданные/размеры — НЕ полный текст документа (§: no-secrets/PII).
"""

import logging
import os
import shutil
import tempfile

from sqlalchemy import select, delete

from ..config import get_settings
from ..database import async_session
from ..models.document import DocumentRecord, DocumentChunk
from . import s3

logger = logging.getLogger("meridian.documents")


# --- извлечение текста по сегментам (page/sheet metadata) ---

def _extract_segments(path: str, ext: str) -> tuple[list[dict], int | None, int | None]:
    """Вернуть (segments, page_count, sheet_count).

    segment = {"text": str, "page_number": int|None, "sheet_name": str|None}
    """
    ext = ext.lower()
    if ext in (".txt", ".md", ".csv"):
        text = _read_text(path)
        return ([{"text": text, "page_number": None, "sheet_name": None}], None, None)
    if ext == ".docx":
        return (_extract_docx(path), None, None)
    if ext == ".xlsx":
        segs = _extract_xlsx(path)
        return (segs, None, len(segs))
    if ext == ".pdf":
        segs = _extract_pdf(path)
        return (segs, len(segs), None)
    raise ValueError(f"Формат {ext} не поддерживается для извлечения текста")


def _read_text(path: str) -> str:
    with open(path, "rb") as f:
        return f.read().decode("utf-8", errors="replace")


def _extract_docx(path: str) -> list[dict]:
    try:
        from docx import Document as DocxDocument
    except ImportError as e:
        raise RuntimeError("python-docx не установлен") from e
    doc = DocxDocument(path)
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    text = "\n\n".join(paragraphs)
    return [{"text": text, "page_number": None, "sheet_name": None}]


def _extract_xlsx(path: str) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError("openpyxl не установлен") from e
    wb = load_workbook(path, read_only=True, data_only=True)
    segments: list[dict] = []
    try:
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                line = "\t".join(cells).rstrip()
                if line.strip():
                    rows.append(line)
            if rows:
                segments.append({
                    "text": f"[Лист: {sheet}]\n" + "\n".join(rows),
                    "page_number": None,
                    "sheet_name": sheet,
                })
    finally:
        wb.close()
    return segments


def _extract_pdf(path: str) -> list[dict]:
    try:
        from PyPDF2 import PdfReader
    except ImportError as e:
        raise RuntimeError("PyPDF2 не установлен") from e
    reader = PdfReader(path)
    segments: list[dict] = []
    for i, page in enumerate(reader.pages):
        text = (page.extract_text() or "").strip()
        if text:
            segments.append({"text": text, "page_number": i + 1, "sheet_name": None})
    if not segments:
        # пустой/сканированный PDF — вернём пустой сегмент, обработчик выдаст ошибку
        return []
    return segments


# --- чанкинг (char-based, с overlap) ---

def chunk_text(text: str, target_chars: int, overlap_chars: int) -> list[str]:
    """Разбить текст на чанки ~target_chars с overlap_chars, по возможности по границам слов/строк."""
    text = (text or "").strip()
    if not text:
        return []
    if len(text) <= target_chars:
        return [text]
    overlap = max(0, min(overlap_chars, target_chars - 1))
    chunks: list[str] = []
    start, n = 0, len(text)
    while start < n:
        end = min(start + target_chars, n)
        if end < n:
            window_start = max(start + 1, end - overlap)
            br = text.rfind("\n", window_start, end)
            if br == -1:
                br = text.rfind(" ", window_start, end)
            if br > start:
                end = br
        chunk = text[start:end].strip()
        if chunk:
            chunks.append(chunk)
        if end >= n:
            break
        start = max(end - overlap, start + 1)
    return chunks


# --- job handler ---

async def handle_document_process(payload: dict) -> None:
    document_id = payload["document_id"]
    settings = get_settings()
    tmpdir: str | None = None
    try:
        async with async_session() as db:
            doc = await db.get(DocumentRecord, document_id)
            if not doc:
                logger.warning("document %s not found", document_id)
                return
            if doc.status == "ready":
                return  # идемпотентность
            doc.status = "processing"
            await db.commit()
            s3_key, ext, original, owner = doc.s3_key, doc.file_ext, doc.original_name, doc.owner_user_id

        if not s3_key:
            raise ValueError("Документ без s3_key")

        tmpdir = tempfile.mkdtemp(prefix="meridian_doc_")
        local = os.path.join(tmpdir, "src" + (ext or ""))
        await s3.download_to(s3_key, local)

        segments, page_count, sheet_count = _extract_segments(local, ext or "")
        full_chars = sum(len(s["text"]) for s in segments)
        if full_chars == 0:
            raise ValueError("Не удалось извлечь текст (пустой или сканированный документ)")

        # чанкинг по сегментам с метаданными
        chunk_rows: list[dict] = []
        idx = 0
        total = 0
        cap = settings.document_max_extract_chars
        for seg in segments:
            for piece in chunk_text(seg["text"], settings.document_chunk_target_chars, settings.document_chunk_overlap_chars):
                if total + len(piece) > cap:
                    logger.warning("document %s: достигнут лимит extract chars, обрезано", document_id)
                    break
                chunk_rows.append({
                    "idx": idx,
                    "text": piece,
                    "page": seg["page_number"],
                    "sheet": seg["sheet_name"],
                    "tokens": len(piece.split()),
                })
                idx += 1
                total += len(piece)
            else:
                continue
            break

        if not chunk_rows:
            raise ValueError("Не удалось сформировать чанки документа")

        # опционально: извлечённый текст целиком в S3
        extracted_key = None
        try:
            extracted_key = s3.object_key(owner, settings.s3_extracted_text_prefix, (original or "doc") + ".txt")
            full_text = "\n\n".join(s["text"] for s in segments)
            await s3.put_bytes(extracted_key, full_text.encode("utf-8"))
        except Exception as e:
            logger.info("document %s: extracted text upload skipped (%s)", document_id, type(e).__name__)
            extracted_key = None

        async with async_session() as db:
            doc = await db.get(DocumentRecord, document_id)
            if not doc:
                return
            await db.execute(delete(DocumentChunk).where(DocumentChunk.document_id == document_id))
            for r in chunk_rows:
                db.add(DocumentChunk(
                    document_id=document_id,
                    chunk_index=r["idx"],
                    text=r["text"],
                    page_number=r["page"],
                    sheet_name=r["sheet"],
                    token_count=r["tokens"],
                ))
            doc.status = "ready"
            doc.page_count = page_count
            doc.sheet_count = sheet_count
            doc.extracted_text_s3_key = extracted_key
            doc.processing_error = None
            await db.commit()
        logger.info("document %s processed: %d chunks, pages=%s sheets=%s",
                    document_id, len(chunk_rows), page_count, sheet_count)
    except Exception as e:
        logger.error("document %s processing failed: %s", document_id, e)
        async with async_session() as db:
            doc = await db.get(DocumentRecord, document_id)
            if doc:
                doc.status = "error"
                doc.processing_error = str(e)[:1000]
                await db.commit()
    finally:
        if tmpdir:
            shutil.rmtree(tmpdir, ignore_errors=True)
